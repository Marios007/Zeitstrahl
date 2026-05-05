import sys
import re
from datetime import datetime, timedelta, time as dtime

import openpyxl
import matplotlib
matplotlib.use("Agg")  # dateibasiertes Backend, kein GUI nötig
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

# Deutsche Farbnamen → Hex
GERMAN_COLORS = {
    "schwarz": "#000000", "rot": "#CC0000", "grün": "#228B22", "gruen": "#228B22",
    "blau": "#1F5FAD", "gelb": "#DAA520", "orange": "#E07000", "lila": "#800080",
    "violett": "#800080", "pink": "#C71585", "rosa": "#FF69B4", "grau": "#808080",
    "weiß": "#FFFFFF", "weiss": "#FFFFFF", "braun": "#8B4513",
    "cyan": "#008B8B", "magenta": "#8B008B",
}

DPI = 150
FIGURE_HEIGHT_IN = 10  # Höhe horizontal / Breite vertikal


def parse_interval(val):
    # Formate: "1h", "30min", "1d", reine Zahl = Minuten
    if val is None:
        return timedelta(hours=1)
    s = str(val).strip().lower()
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*(h|min|m|d|)?", s)
    if not m:
        return timedelta(hours=1)
    num = float(m.group(1))
    unit = m.group(2)
    if unit == "d":
        return timedelta(days=num)
    if unit in ("h", ""):
        return timedelta(hours=num)
    return timedelta(minutes=num)


def resolve_color(raw):
    if raw is None:
        return "#000000"
    s = str(raw).strip().lower()
    if s in GERMAN_COLORS:
        return GERMAN_COLORS[s]
    return raw


def parse_cell_datetime(date_val, time_val):
    # Kombiniert Datums- und Zeitzelle; Excel liefert Zeiten manchmal als datetime
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        base = date_val.date()
    else:
        return None
    if time_val is None:
        t = dtime(0, 0)
    elif isinstance(time_val, str):
        t = datetime.strptime(time_val.strip(), "%H:%M").time()
    elif isinstance(time_val, datetime):
        t = time_val.time()
    else:
        t = time_val
    return datetime.combine(base, t)


def load_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # I2: Tick-Intervall | I19: Projektname | J19: Ausrichtung | K19: Dateiformat
    # N2/O2: Startdatum/-zeit | P2/Q2: Enddatum/-zeit
    # R2: Uhrzeit im Label | S2: Schriftgröße (klein/normal/groß)
    interval = parse_interval(ws["I2"].value)
    date_range = (
        parse_cell_datetime(ws["N2"].value, ws["O2"].value),
        parse_cell_datetime(ws["P2"].value, ws["Q2"].value),
    )
    show_time_in_label = str(ws["R2"].value).strip().lower() == "ja" if ws["R2"].value is not None else False
    font_offset = {"klein": -2, "groß": 2, "gross": 2}.get(
        str(ws["S2"].value).strip().lower() if ws["S2"].value is not None else "", 0
    )
    project_name = str(ws["I19"].value).strip() if ws["I19"].value is not None else ""
    orientation = str(ws["J19"].value).strip().lower() if ws["J19"].value is not None else "beide"
    file_format = str(ws["K19"].value).strip().lower() if ws["K19"].value is not None else "beide"

    # Kategorien aus Spalten J (Name), K (Farbe), L (Sichtbar), M (Stem-Länge)
    color_map = {}
    visible_cats = set()
    stem_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cat_key = row[9]
        if cat_key is not None:
            key = str(cat_key).strip()
            color_map[key] = resolve_color(row[10])
            if row[11] is None or str(row[11]).strip().lower() != "nein":
                visible_cats.add(key)
            stem_map[key] = float(row[12]) if row[12] is not None else 2.5

    # Events aus Spalten A–F
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        datum, uhrzeit, quelle, kategorie, text, wichtig_val = row[:6]

        if datum is None or not isinstance(datum, datetime):
            continue

        if uhrzeit is not None:
            if isinstance(uhrzeit, str):
                t = datetime.strptime(uhrzeit.strip(), "%H:%M").time()
            elif isinstance(uhrzeit, datetime):
                t = uhrzeit.time()
            else:
                t = uhrzeit
        else:
            t = dtime(0, 0)

        dt = datetime.combine(datum.date(), t)
        cat_str = str(kategorie).strip() if kategorie is not None else ""
        if cat_str not in visible_cats:
            continue

        label = "_".join(str(p) for p in [quelle, kategorie, text] if p is not None)
        if show_time_in_label:
            label = f"{label} ({dt.strftime('%H:%M')})"
        wichtig = str(wichtig_val).strip().lower() == "ja" if wichtig_val is not None else False

        events.append({
            "dt": dt,
            "label": label,
            "color": color_map.get(cat_str, "#000000"),
            "stem": stem_map.get(cat_str, 2.5),
            "wichtig": wichtig,
        })

    events.sort(key=lambda e: e["dt"])
    return events, interval, date_range, font_offset, project_name, orientation, file_format


def _tick_positions(t_min, t_max, tick_interval):
    positions = []
    td = t_min
    while td <= t_max + tick_interval:
        positions.append(td)
        td += tick_interval
    return positions


def _apply_date_range(events, date_range):
    start_dt, end_dt = date_range
    if start_dt is not None:
        events = [e for e in events if e["dt"] >= start_dt]
    if end_dt is not None:
        events = [e for e in events if e["dt"] <= end_dt]
    return events


def _wichtig_style(wichtig, font_offset=0):
    # Visuelle Parameter: wichtige Events sind voller und größer
    return {
        "lw": 2.0 if wichtig else 1.2,
        "alpha": 0.4,
        "fsize": (9 if wichtig else 7) + font_offset,
    }


def build_figure(events, interval, date_range=(None, None), font_offset=0, project_name=""):
    if not events:
        print("Keine Ereignisse gefunden.")
        sys.exit(1)

    events = _apply_date_range(events, date_range)
    if not events:
        print("Keine Ereignisse im angegebenen Zeitraum gefunden.")
        sys.exit(1)

    start_dt, end_dt = date_range
    t_min = start_dt if start_dt is not None else events[0]["dt"]
    t_max = end_dt if end_dt is not None else events[-1]["dt"]
    total_seconds = max((t_max - t_min).total_seconds(), 60)

    tick_interval = timedelta(hours=1) if interval == timedelta(days=1) else interval
    num_ticks = int(total_seconds / tick_interval.total_seconds()) + 2

    # Breite: basiert auf Tick-Anzahl, nicht Event-Anzahl
    fig_width_in = max(1000, num_ticks * 80) / DPI
    max_stem = max(ev.get("stem", 2.5) for ev in events)
    y_margin = max_stem + 1.5
    padding = tick_interval * 0.8

    fig, ax = plt.subplots(figsize=(fig_width_in, FIGURE_HEIGHT_IN))
    ax.set_xlim(t_min - padding, t_max + padding)
    ax.set_ylim(-y_margin, y_margin)
    ax.axis("off")

    AXIS_COLOR = "#2c3e50"
    TICK_COLOR = "#7f8c8d"
    GRID_COLOR = "#ecf0f1"

    tick_positions = _tick_positions(t_min, t_max, tick_interval)

    for td in tick_positions:
        x = mdates.date2num(td)
        ax.vlines(x, -y_margin, y_margin, color=GRID_COLOR, linewidth=0.5, zorder=0)

    ax.axhline(0, color=AXIS_COLOR, linewidth=2, zorder=1)
    ax.annotate("", xy=(mdates.date2num(t_max + padding * 0.9), 0),
                xytext=(mdates.date2num(t_max + padding * 0.5), 0),
                arrowprops=dict(arrowstyle="->", color=AXIS_COLOR, lw=1.5))

    for td in tick_positions:
        x = mdates.date2num(td)
        ax.vlines(x, -0.25, 0.25, color=TICK_COLOR, linewidth=1.5, zorder=2)
        # Datum unterhalb, Uhrzeit oberhalb der Achse
        ax.text(x, -0.35, td.strftime("%d.%m.%Y"), ha="center", va="top",
                fontsize=7, color="#000000", rotation=90)
        if not (td.hour == 0 and td.minute == 0 and td.second == 0):
            ax.text(x, 0.35, td.strftime("%H:%M"), ha="center", va="bottom",
                    fontsize=7, color="#000000", rotation=90)

    for i, ev in enumerate(events):
        x = mdates.date2num(ev["dt"])
        color = ev["color"]
        stem_length = ev.get("stem", 2.5)
        sign = 1 if i % 2 == 0 else -1  # abwechselnd oben/unten
        wichtig = ev.get("wichtig", False)
        style = _wichtig_style(wichtig, font_offset)

        ax.plot(x, 0, "o", color=color, markersize=6, zorder=3)
        y_end = sign * stem_length
        ax.plot([x, x], [0, y_end], color=color, linewidth=style["lw"],
                alpha=style["alpha"], zorder=2)
        if wichtig:
            ax.plot(x, y_end, "*", color=color, markersize=14, zorder=4)

        text_offset = 0.4 if wichtig else 0.15
        y_text = y_end + sign * text_offset
        va = "bottom" if sign > 0 else "top"
        text_kwargs = dict(ha="center", va=va, fontsize=style["fsize"],
                           color=color, rotation=90, fontweight="bold")
        if wichtig:
            text_kwargs["bbox"] = dict(boxstyle="round,pad=0.3", facecolor="white",
                                       edgecolor=color, linewidth=1.2, alpha=0.9)
        ax.text(x, y_text, ev["label"], **text_kwargs)

    meta = f"Projekt: {project_name}  |  Intervall: {_interval_str(interval)}  |  Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    fig.text(0.01, 0.01, meta, fontsize=6, color="#aaaaaa", va="bottom")
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")
    plt.tight_layout(pad=0.5)
    return fig


def build_figure_vertical(events, interval, date_range=(None, None), font_offset=0, project_name=""):
    if not events:
        print("Keine Ereignisse gefunden.")
        sys.exit(1)

    events = _apply_date_range(events, date_range)
    if not events:
        print("Keine Ereignisse im angegebenen Zeitraum gefunden.")
        sys.exit(1)

    start_dt, end_dt = date_range
    t_min = start_dt if start_dt is not None else events[0]["dt"]
    t_max = end_dt if end_dt is not None else events[-1]["dt"]
    total_seconds = max((t_max - t_min).total_seconds(), 60)

    tick_interval = timedelta(hours=1) if interval == timedelta(days=1) else interval
    num_ticks = int(total_seconds / tick_interval.total_seconds()) + 2

    fig_height_in = max(1000, num_ticks * 80) / DPI
    fig_width_in = FIGURE_HEIGHT_IN
    max_stem = max(ev.get("stem", 2.5) for ev in events)
    x_margin = max_stem + 2.5
    padding = tick_interval * 0.8

    fig, ax = plt.subplots(figsize=(fig_width_in, fig_height_in))
    ax.set_xlim(-x_margin, x_margin)
    # Y-Achse invertiert: ältestes oben, neuestes unten
    ax.set_ylim(mdates.date2num(t_max + padding), mdates.date2num(t_min - padding))
    ax.axis("off")

    AXIS_COLOR = "#2c3e50"
    TICK_COLOR = "#7f8c8d"
    GRID_COLOR = "#ecf0f1"

    tick_positions = _tick_positions(t_min, t_max, tick_interval)

    for td in tick_positions:
        y = mdates.date2num(td)
        ax.hlines(y, -x_margin, x_margin, color=GRID_COLOR, linewidth=0.5, zorder=0)

    ax.axvline(0, color=AXIS_COLOR, linewidth=2, zorder=1)
    ax.annotate("", xy=(0, mdates.date2num(t_max + padding * 0.9)),
                xytext=(0, mdates.date2num(t_max + padding * 0.5)),
                arrowprops=dict(arrowstyle="->", color=AXIS_COLOR, lw=1.5))

    for td in tick_positions:
        y = mdates.date2num(td)
        ax.hlines(y, -0.25, 0.25, color=TICK_COLOR, linewidth=1.5, zorder=2)
        # Datum links, Uhrzeit rechts der Achse
        ax.text(-0.35, y, td.strftime("%d.%m.%Y"), ha="right", va="center",
                fontsize=7, color="#000000", rotation=0)
        if not (td.hour == 0 and td.minute == 0 and td.second == 0):
            ax.text(0.35, y, td.strftime("%H:%M"), ha="left", va="center",
                    fontsize=7, color="#000000", rotation=0)

    for i, ev in enumerate(events):
        y = mdates.date2num(ev["dt"])
        color = ev["color"]
        stem_length = ev.get("stem", 2.5)
        sign = 1 if i % 2 == 0 else -1  # abwechselnd links/rechts
        wichtig = ev.get("wichtig", False)
        style = _wichtig_style(wichtig, font_offset)

        ax.plot(0, y, "o", color=color, markersize=6, zorder=3)
        x_end = sign * stem_length
        ax.plot([0, x_end], [y, y], color=color, linewidth=style["lw"],
                alpha=style["alpha"], zorder=2)
        if wichtig:
            ax.plot(x_end, y, "*", color=color, markersize=14, zorder=4)

        text_offset = 0.4 if wichtig else 0.15
        x_text = x_end + sign * text_offset
        ha = "left" if sign > 0 else "right"
        text_kwargs = dict(ha=ha, va="center", fontsize=style["fsize"],
                           color=color, fontweight="bold")
        if wichtig:
            text_kwargs["bbox"] = dict(boxstyle="round,pad=0.3", facecolor="white",
                                       edgecolor=color, linewidth=1.2, alpha=0.9)
        ax.text(x_text, y, ev["label"], **text_kwargs)

    meta = f"Projekt: {project_name}  |  Intervall: {_interval_str(interval)}  |  Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    fig.text(0.01, 0.01, meta, fontsize=6, color="#aaaaaa", va="bottom")
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")
    plt.tight_layout(pad=0.5)
    return fig


def _interval_str(td):
    total = int(td.total_seconds())
    if total % 86400 == 0:
        return f"{total // 86400}d"
    if total % 3600 == 0:
        return f"{total // 3600}h"
    return f"{total // 60}min"


def main():
    import os
    path = "Zeitstrahl.xlsx"
    events, interval, date_range, font_offset, project_name, orientation, file_format = load_excel(path)
    print(f"{len(events)} Ereignisse geladen, Intervall: {interval}")

    os.makedirs("output", exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = _interval_str(interval)

    figures = []
    beide = {"beide", "beides"}
    if orientation in {"horizontal"} | beide:
        figures.append((build_figure(events, interval, date_range, font_offset, project_name), f"Zeitstrahl_{suffix}"))
    if orientation in {"vertikal"} | beide:
        figures.append((build_figure_vertical(events, interval, date_range, font_offset, project_name), f"Zeitstrahl_vertikal_{suffix}"))

    for fig, name in figures:
        base = os.path.join("output", f"{ts}_{name}")
        if file_format in {"png"} | beide:
            fig.savefig(f"{base}.png", dpi=DPI, bbox_inches="tight", facecolor="white")
            print(f"PNG gespeichert: {base}.png")
        if file_format in {"pdf"} | beide:
            fig.savefig(f"{base}.pdf", bbox_inches="tight", facecolor="white")
            print(f"PDF gespeichert: {base}.pdf")
        plt.close(fig)


if __name__ == "__main__":
    main()
